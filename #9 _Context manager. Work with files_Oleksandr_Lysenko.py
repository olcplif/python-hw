import pickle

from openpyxl import load_workbook

# Task_1: Subtitle file processing
dct = {}
keys_list = []
value_list = []

with open("files_contextmanagers/task1.txt", "r") as file:
    context_list = list(file.readlines())

# 1.1. Make Dict
for i in range(0, len(context_list), 2):
    keys_list.append(context_list[i])

for i in range(1, len(context_list), 2):
    value_list.append(context_list[i])

dct = dict.fromkeys(keys_list)

i = 0
for key in dct:
    dct.update({key: value_list[i]})
    i += 1

print(dct)

# 1.2. Make file

new_context_list = []

for i in range(1, len(context_list), 2):
    new_context_list.append(context_list[i].replace("\n", ""))

lst_to_str = ' '.join([str(elem) for elem in new_context_list])
print(lst_to_str)

with open("files_contextmanagers/task1-2.txt", "w") as file:
    file.write(lst_to_str)


# Task_2: Read list from file and get average values

with open("files_contextmanagers/task2", "rb") as file:
    byte_lst = file.read()

lst = pickle.loads(byte_lst)
print(lst)
mean_lst = int(sum(lst) / len(lst))
print(mean_lst)


# Task_3: Context manager for excel files

class OpenXlsx:

    def __init__(self, path):
        self.file = load_workbook(path)

    def __enter__(self):
        return self.file

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.file.close()
        elif exc_type is Exception:
            print(f'An error(s) occurred while working with the file: {exc_val}')
        pass


path_to_file = 'files_contextmanagers/task3.xlsx'
with OpenXlsx(path_to_file) as file:
    print(f'List pages of file: {file.sheetnames}')

    file.create_sheet("New page", 3)
    print(f'List pages of file after adding: {file.sheetnames}')

    # Get workbook active sheet object
    work_sheet = file.active
    print(f"Active sheet: {work_sheet.title}")

    # Cell object is created and added value
    work_sheet['A1'] = "ID"
    work_sheet['B1'] = "Name"
    work_sheet['C1'] = "Age"

    work_sheet['A2'] = "1"
    work_sheet['B2'] = "Ustym"
    work_sheet['C2'] = 5

    work_sheet['A3'] = "2"
    work_sheet['B3'] = "Mykyta"
    work_sheet['C3'] = 2

    work_sheet['B5'] = "Average age:"
    work_sheet['C5'] = (work_sheet['C2'].value + work_sheet['C3'].value) / 2
    work_sheet['B6'] = "Count people:"
    work_sheet['C6'] = "=COUNT(C2;C3)"

    # If exception was raised - file NOT save!
    # raise Exception

    file.save(path_to_file)

print("File closed!")
